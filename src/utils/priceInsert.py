import openpyxl
import psycopg2
import pandas as pd

# 读取 Excel 文件
file_path = '../Price.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# 从Excel中提取数据
data = []
start_stations = [cell.value for row in sheet.iter_rows(min_row=4, max_row=376, min_col=3, max_col=3) for cell in row]
end_stations = [cell.value for cell in sheet[3][3:376]]
prices = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=4, max_row=376, min_col=4, max_col=376)]

# 连接到 PostgreSQL 数据库
conn = psycopg2.connect(
    dbname="project",
    user="checker",
    password="123456",
    host="localhost",
)
cursor = conn.cursor()

# 执行参数化的INSERT查询
for i, start_station in enumerate(start_stations):
    for j, end_station in enumerate(end_stations):
        price = prices[i][j]
        if price is not None:
            # 获取起始站点和终点站点的ID
            cursor.execute("SELECT id FROM station WHERE chinese_name = %s", (start_station,))
            start_station_id = cursor.fetchone()[0]
            cursor.execute("SELECT id FROM station WHERE chinese_name = %s", (end_station,))
            end_station_id = cursor.fetchone()[0]

            # 插入数据
            cursor.execute("INSERT INTO price (start_station, end_station, price) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
                           (start_station_id, end_station_id, price))

        # 提交更改并关闭连接
conn.commit()
cursor.close()
conn.close()